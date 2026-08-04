# -*- coding: utf-8 -*-
"""
《K線西遊記・台指期貨轉檔系統》

固定輸入／輸出規格：
- 期交所原始資料固定 19 欄。
- 僅處理契約 TX，排除跨月價差列。
- 每個交易日選擇該日仍存在的最小月契約作為近月。
- 結算日仍使用當月契約；下一交易日才依原始資料自然換月。
- 全日 K：夜盤開、日盤收、日夜最高最低、日夜成交量相加。
- 結算價與未沖銷契約數只取日盤資料。
- 主檔固定為 master/台指近全.xlsx，並另產出台指近全YYYYMMDD.xlsx。

重要：
不得用 split(',') 解析期交所 CSV。必須遵守 CSV 引號與換行規則，
解析失敗時立即停止，禁止用錯位資料覆蓋主檔。
"""
from __future__ import annotations

import argparse
import csv
import os
import re
from io import StringIO
from pathlib import Path
from typing import Iterable, List, Optional

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

HEADERS19: List[str] = [
    "交易日期", "契約", "到期月份(週別)", "開盤價", "最高價", "最低價", "收盤價",
    "漲跌價", "漲跌%", "成交量", "結算價", "未沖銷契約數", "最後最佳買價", "最後最佳賣價",
    "歷史最高價", "歷史最低價", "是否因訊息面暫停交易", "交易時段", "價差對單式委託成交量",
]

ENCODINGS = ("utf-8-sig", "utf-8", "cp950", "big5")
DAY_TAGS = {"一般", "@"}
NIGHT_TAGS = {"盤後", "L"}
NUMERIC_COLUMNS = [
    "開盤價", "最高價", "最低價", "收盤價", "漲跌價", "成交量", "結算價",
    "未沖銷契約數", "最後最佳買價", "最後最佳賣價", "歷史最高價", "歷史最低價",
    "價差對單式委託成交量",
]


def _read_text_any(path: str) -> str:
    last_error: Optional[Exception] = None
    for encoding in ENCODINGS:
        try:
            return Path(path).read_text(encoding=encoding)
        except Exception as exc:  # pragma: no cover - only used for fallback encodings
            last_error = exc
    raise RuntimeError(f"讀檔失敗，無法識別編碼：{path}；{last_error}")


def _is_date_token(value: object) -> bool:
    text = str(value).strip()
    return bool(
        re.fullmatch(r"\d{4}[-/]\d{1,2}[-/]\d{1,2}", text)
        or re.fullmatch(r"\d{8}", text)
        or re.fullmatch(r"\d{7}", text)
    )


def _normalize_date_value(value: object) -> int:
    text = str(value).strip()
    digits = re.sub(r"\D", "", text)
    if len(digits) == 8:
        return int(digits)
    if len(digits) == 7:
        # 民國年月日，例如 1150804。
        roc_year = int(digits[:3])
        return int(f"{roc_year + 1911:04d}{digits[3:]}")
    raise ValueError(f"無效交易日期：{value!r}")


def _normalize_date_series(series: pd.Series) -> pd.Series:
    return series.map(_normalize_date_value).astype("int64")


def _clean_row(row: Iterable[object]) -> List[str]:
    return [str(cell).replace("\ufeff", "").strip() for cell in row]


def read_csv_strict_19_multiline(path: str) -> pd.DataFrame:
    """以標準 CSV 規則解析，並相容真正被拆行的 19 欄記錄。

    csv.reader 會正確處理引號內逗號與引號內換行。若來源真的把一筆資料
    無引號拆成多個實體列，才會把不足 19 欄的相鄰列接回；任何超過 19 欄
    或接回後仍不足 19 欄的資料一律拒絕，避免靜默欄位錯位。
    """
    text = _read_text_any(path)
    parsed_rows = [_clean_row(row) for row in csv.reader(StringIO(text, newline=""))]
    parsed_rows = [row for row in parsed_rows if any(cell != "" for cell in row)]
    if not parsed_rows:
        raise ValueError(f"未解析到任何資料：{path}")

    records: List[List[str]] = []
    buffer: List[str] = []

    for row_number, row in enumerate(parsed_rows, start=1):
        if row == HEADERS19:
            if buffer:
                raise ValueError(f"第 {row_number} 列表頭前仍有未完成資料：{path}")
            continue

        if len(row) > 19:
            raise ValueError(
                f"第 {row_number} 列解析為 {len(row)} 欄，應為 19 欄；"
                "來源可能有未加引號的逗號，已停止以防主檔被錯誤覆蓋。"
            )

        if buffer and row and _is_date_token(row[0]):
            raise ValueError(
                f"第 {row_number} 列已開始新日期，但上一筆只有 {len(buffer)} 欄；"
                "禁止自動補空值造成欄位錯位。"
            )

        buffer.extend(row)
        if len(buffer) == 19:
            records.append(buffer)
            buffer = []
        elif len(buffer) > 19:
            raise ValueError(f"第 {row_number} 列合併後超過 19 欄：{path}")

    if buffer:
        raise ValueError(f"檔尾仍有未完成資料，共 {len(buffer)} 欄：{path}")
    if not records:
        raise ValueError(f"沒有完整 19 欄資料列：{path}")

    frame = pd.DataFrame(records, columns=HEADERS19)
    invalid_dates = ~frame["交易日期"].map(_is_date_token)
    if invalid_dates.any():
        examples = frame.loc[invalid_dates, "交易日期"].head(3).tolist()
        raise ValueError(f"交易日期欄錯位或格式錯誤：{examples}")
    return frame


def _extract_contract_month(series: pd.Series) -> pd.Series:
    return series.astype(str).str.strip().str.extract(r"^(\d{6})$")[0]


def _pick_near_month_for_day(group: pd.DataFrame) -> int:
    """選擇當日存在的最小月契約。

    結算日仍屬當月，不能因結算價為 0 就提前切換次月。結算後下一交易日，
    到期契約不再出現在原始資料中，最小月契約會自然成為次月。
    """
    months = pd.to_numeric(group["_contract_month"], errors="coerce").dropna().astype(int)
    if months.empty:
        raise ValueError("當日找不到六碼月契約")
    return int(months.min())


def _last_non_null(series: pd.Series):
    values = series.dropna()
    return values.iloc[-1] if not values.empty else np.nan


def _first_non_null(series: pd.Series):
    values = series.dropna()
    return values.iloc[0] if not values.empty else np.nan


def make_near_full_day(df19: pd.DataFrame) -> pd.DataFrame:
    df = df19.copy()
    df["交易日期"] = _normalize_date_series(df["交易日期"])
    df["契約"] = df["契約"].astype(str).str.strip().str.upper()
    df["到期月份(週別)"] = df["到期月份(週別)"].astype(str).str.strip()
    df["交易時段"] = df["交易時段"].astype(str).str.strip()

    # 僅台指月契約；跨月價差與週契約不得混入近月主檔。
    df = df[df["契約"] == "TX"].copy()
    df = df[~df["到期月份(週別)"].str.contains("/", na=False)].copy()
    df["_contract_month"] = _extract_contract_month(df["到期月份(週別)"])
    df = df[df["_contract_month"].notna()].copy()
    if df.empty:
        raise ValueError("原始檔沒有可用的 TX 六碼月契約資料")

    unknown_sessions = sorted(set(df["交易時段"]) - DAY_TAGS - NIGHT_TAGS)
    if unknown_sessions:
        raise ValueError(f"發現未知交易時段：{unknown_sessions}")

    for column in NUMERIC_COLUMNS:
        df[column] = pd.to_numeric(
            df[column].astype(str).str.replace(",", "", regex=False).str.replace("%", "", regex=False),
            errors="coerce",
        )

    session_order = {"盤後": 0, "L": 0, "一般": 1, "@": 1}
    df["_session_order"] = df["交易時段"].map(session_order)
    df["_source_order"] = np.arange(len(df))
    df = df.sort_values(["交易日期", "_session_order", "_source_order"])

    records: List[dict] = []
    for trade_date, day_group in df.groupby("交易日期", sort=True):
        near_month = _pick_near_month_for_day(day_group)
        near_group = day_group[day_group["_contract_month"].astype(int) == near_month].copy()
        day_part = near_group[near_group["交易時段"].isin(DAY_TAGS)]
        night_part = near_group[near_group["交易時段"].isin(NIGHT_TAGS)]

        open_price = _first_non_null(night_part["開盤價"])
        if pd.isna(open_price):
            open_price = _first_non_null(day_part["開盤價"])

        close_price = _last_non_null(day_part["收盤價"])
        if pd.isna(close_price):
            close_price = _last_non_null(night_part["收盤價"])

        base_row = day_part.iloc[-1] if not day_part.empty else near_group.iloc[-1]
        change_pct = str(base_row.get("漲跌%", "")).strip()
        if change_pct.lower() in {"", "nan", "none"}:
            change_pct = ""
        elif not change_pct.endswith("%"):
            change_pct = f"{change_pct}%"

        record = {
            "交易日期": int(trade_date),
            "契約": "TX",
            "到期月份(週別)": str(near_month),
            "開盤價": open_price,
            "最高價": near_group["最高價"].max(skipna=True),
            "最低價": near_group["最低價"].min(skipna=True),
            "收盤價": close_price,
            "漲跌價": _last_non_null(day_part["漲跌價"]) if not day_part.empty else base_row.get("漲跌價", np.nan),
            "漲跌%": change_pct,
            "成交量": near_group["成交量"].fillna(0).sum(),
            # 結算與未平倉只採日盤；夜盤不得覆蓋。
            "結算價": _last_non_null(day_part["結算價"]) if not day_part.empty else np.nan,
            "未沖銷契約數": _last_non_null(day_part["未沖銷契約數"]) if not day_part.empty else np.nan,
            "最後最佳買價": base_row.get("最後最佳買價", np.nan),
            "最後最佳賣價": base_row.get("最後最佳賣價", np.nan),
            "歷史最高價": base_row.get("歷史最高價", np.nan),
            "歷史最低價": base_row.get("歷史最低價", np.nan),
            "是否因訊息面暫停交易": base_row.get("是否因訊息面暫停交易", ""),
            "交易時段": "全日",
            "價差對單式委託成交量": base_row.get("價差對單式委託成交量", np.nan),
        }
        records.append(record)

    result = pd.DataFrame(records, columns=HEADERS19)
    validate_near_full(result)
    return result


def validate_near_full(frame: pd.DataFrame) -> None:
    if list(frame.columns) != HEADERS19:
        raise ValueError("輸出欄位不是固定 19 欄或順序錯誤")
    if frame.empty:
        raise ValueError("近月全日資料為空")
    if frame["交易日期"].duplicated().any():
        raise ValueError("同一交易日期出現多筆近月資料")
    if not frame["交易日期"].is_monotonic_increasing:
        raise ValueError("交易日期未依升冪排列")
    if not frame["到期月份(週別)"].astype(str).str.fullmatch(r"\d{6}").all():
        raise ValueError("近月契約不是六碼月份")

    numeric = frame[["開盤價", "最高價", "最低價", "收盤價"]].apply(pd.to_numeric, errors="coerce")
    if numeric.isna().any().any():
        raise ValueError("OHLC 有空值或非數字")
    invalid_ohlc = (
        (numeric["最高價"] < numeric[["開盤價", "收盤價", "最低價"]].max(axis=1))
        | (numeric["最低價"] > numeric[["開盤價", "收盤價", "最高價"]].min(axis=1))
    )
    if invalid_ohlc.any():
        dates = frame.loc[invalid_ohlc, "交易日期"].tolist()
        raise ValueError(f"OHLC 邏輯錯誤，日期：{dates}")
    if (pd.to_numeric(frame["成交量"], errors="coerce") < 0).any():
        raise ValueError("成交量不可為負數")


def _freeze_first_row_and_autowidth(path: str, sheet_name: Optional[str] = None) -> None:
    workbook = load_workbook(path)
    worksheet = workbook.active if not sheet_name or sheet_name not in workbook.sheetnames else workbook[sheet_name]
    worksheet.freeze_panes = "A2"
    for index, cell in enumerate(worksheet[1], start=1):
        header = "" if cell.value is None else str(cell.value)
        worksheet.column_dimensions[get_column_letter(index)].width = max(12, min(30, len(header) + 2))
    workbook.save(path)


def _standard_master_name(date_yyyymmdd: int) -> str:
    return f"台指近全{date_yyyymmdd}.xlsx"


def _load_master(path: Optional[str]) -> pd.DataFrame:
    if not path or not os.path.exists(path):
        return pd.DataFrame(columns=HEADERS19)
    master = pd.read_excel(path)
    missing = [column for column in HEADERS19 if column not in master.columns]
    if missing:
        raise ValueError(f"既有主檔缺少欄位：{missing}")
    master = master[HEADERS19].copy()
    if not master.empty:
        master["交易日期"] = _normalize_date_series(master["交易日期"])
        master = master.drop_duplicates(subset=["交易日期"], keep="last").sort_values("交易日期")
    return master


def _write_excel(frame: pd.DataFrame, path: str, sheet_name: str) -> None:
    Path(path).parent.mkdir(parents=True, exist_ok=True)
    frame.to_excel(path, index=False, sheet_name=sheet_name)
    _freeze_first_row_and_autowidth(path, sheet_name)


def run(input_csvs: List[str], master_path: Optional[str], output: Optional[str]) -> List[str]:
    master = _load_master(master_path)
    generated: List[str] = []
    latest_date: Optional[int] = None

    for index, csv_path in enumerate(sorted(input_csvs)):
        source = read_csv_strict_19_multiline(csv_path)
        near = make_near_full_day(source)

        if output and len(input_csvs) > 1:
            raise ValueError("批次處理多個 CSV 時不可指定單一 --output")
        near_output = output or str(Path(csv_path).with_suffix("")) + "_near_full_v86.xlsx"
        if near_output.lower().endswith(".csv"):
            near.to_csv(near_output, index=False, encoding="utf-8-sig")
        else:
            _write_excel(near, near_output, "台指近全")
        generated.append(near_output)

        new_dates = near["交易日期"].astype(int).tolist()
        master = master[~master["交易日期"].isin(new_dates)].copy()
        master = pd.concat([master, near], ignore_index=True)
        master = master.drop_duplicates(subset=["交易日期"], keep="last").sort_values("交易日期").reset_index(drop=True)
        validate_near_full(master)
        latest_date = max(new_dates) if latest_date is None else max(latest_date, max(new_dates))

    if latest_date is None:
        raise ValueError("沒有產出任何 TX 近月全日資料")

    output_dir = Path(master_path).parent if master_path else Path(input_csvs[-1]).parent
    dated_master = str(output_dir / _standard_master_name(latest_date))
    _write_excel(master, dated_master, "台指近全")
    generated.append(dated_master)

    if master_path:
        _write_excel(master, master_path, "台指近全")
        generated.append(master_path)

    return generated


def main() -> None:
    parser = argparse.ArgumentParser(description="台指期 TX 近月全日 K 轉檔（固定 19 欄）")
    parser.add_argument("input_csv", nargs="+", help="期交所原始 19 欄 CSV，可輸入多檔")
    parser.add_argument("--master", default=None, help="固定主檔台指近全.xlsx")
    parser.add_argument("-o", "--output", default=None, help="單檔近月輸出 xlsx/csv")
    args = parser.parse_args()

    try:
        outputs = run(args.input_csv, args.master, args.output)
    except Exception as exc:
        raise SystemExit(f"TAIFEX 轉檔失敗：{exc}") from exc

    for path in outputs:
        print(path)


if __name__ == "__main__":
    main()
